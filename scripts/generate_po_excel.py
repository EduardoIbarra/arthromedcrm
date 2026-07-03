import sys
import json
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

def generate_po(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
        
    date_str = config.get('date', '')
    numero_orden = config.get('numero_orden', '')
    items = config.get('items', [])
    output_path = config.get('output_path', '')
    
    # Format date: month day year -> MM/DD/YYYY
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        formatted_date = dt.strftime('%m/%d/%Y')
    except Exception as e:
        formatted_date = datetime.now().strftime('%m/%d/%Y')
        
    wb = openpyxl.load_workbook('/Users/ed/Downloads/PO BONSS Medical actualizado.xlsx', data_only=False)
    ws = wb.active
    
    # Update date in E3
    ws.cell(row=3, column=5).value = f"Date {formatted_date}"
    
    # Update Pre-Order title in A4 to include the order number
    ws.cell(row=4, column=1).value = f"Pre-Order {numero_orden}"
    
    # Collect items that we want to keep
    # Items from config have: model, code, quantity
    ordered_items = {}
    for item in items:
        m = str(item.get('model', '')).strip().lower()
        c = str(item.get('code', '')).strip().lower()
        ordered_items[(m, c)] = item.get('quantity', 0)
        
    # We will iterate from row 142 down to 7 (the products range)
    # If a product is ordered, we update the quantity in Col 5 (Qty).
    # If not, we delete the row.
    rows_to_delete = []
    
    for r in range(142, 6, -1):
        cell_model = ws.cell(row=r, column=2).value
        cell_code = ws.cell(row=r, column=3).value
        
        m_key = str(cell_model).strip().lower() if cell_model is not None else ""
        c_key = str(cell_code).strip().lower() if cell_code is not None else ""
        
        # Try to find match
        qty = None
        # Match by both model and code
        if (m_key, c_key) in ordered_items:
            qty = ordered_items[(m_key, c_key)]
        # Match by model only
        elif m_key and any(k[0] == m_key for k in ordered_items):
            matching_keys = [k for k in ordered_items if k[0] == m_key]
            qty = ordered_items[matching_keys[0]]
        # Match by code only
        elif c_key and any(k[1] == c_key for k in ordered_items):
            matching_keys = [k for k in ordered_items if k[1] == c_key]
            qty = ordered_items[matching_keys[0]]
            
        if qty is not None and qty > 0:
            ws.cell(row=r, column=5).value = qty
        else:
            rows_to_delete.append(r)
            
    # Delete non-ordered rows in reverse order (which is already sorted descending in rows_to_delete)
    for r in rows_to_delete:
        ws.delete_rows(r, 1)
        
    # Find the new total row
    total_row = None
    for r in range(7, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == 'TOTAL':
            total_row = r
            break
            
    if total_row:
        # Update sum formula to match the new range: E7 to E{total_row-1}
        ws.cell(row=total_row, column=5).value = f"=SUM(E7:E{total_row - 1})"
        
    wb.save(output_path)
    print(f"Generated PO Excel at {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 generate_po_excel.py <config_json_path>")
        sys.exit(1)
    generate_po(sys.argv[1])
